from openpyxl.worksheet import worksheet
from openpyxl import Workbook
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Font
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension
from openpyxl.utils import get_column_letter
from analyst import *

class ExcelReport:
    """
    Класса для создания отчета в виде Excel таблиц
    :param border: границы ячеек в таблице
    :type border: Border
    :param font: шрифт текста в таблице
    :type font: Font
    """
    def __init__(self, side, font):
        """
        Инициализирует объект класса ExcelReport
        :param side: тип границ для ячеек таблицы
        :type side: Side
        :param font: шрифт текста в отчете
        :type font: Font
        """
        self.border = Border(left=side, top=side, right=side, bottom=side)
        self.font = font

    def generate_excel(self, data: Data):
        """
        Генерирует excel-файл со статистикой собранной из входных данных
        :param dicts: вхожные данные
        :type dicts: list
        :return: сгенерированный excel-файл со статистикой в виде таблиц
        """
        wb = Workbook()
        self.generate_years_sheet(wb, data.salary_by_years, data.count_by_years)
        self.generate_areas_sheet(wb, data.salary_by_cities, data.count_by_cities)
        ws = wb.worksheets[0]
        wb.remove(ws)
        wb.save("android_analyst.xlsx")

    def generate_years_sheet(self, wb: Workbook, salaries: dict, counts: dict) -> worksheet.Worksheet:
        """
        Генирирует лист в excel-книге со статистикой по годам
        :param wb: excel-книга
        :type wb: Workbook
        :param salaries: словарь средних зарплат по годам
        :type salaries: dict
        :param prof_salaries: словарь средних зарплат по годам для введенной профессии
        :type prof_salaries: dict
        :param counts: словарь количества вакансий по годам
        :type counts: dict
        :param prof_counts: словарь количества вакансий по годам для введенной профессии
        :type prof_counts: dict
        :return: excel-лист со статистикой по годам
        """
        years_sheet = wb.create_sheet("Статистика по годам")
        years_sheet['A1'] = 'Год'
        years_sheet['B1'] = 'Средняя зарплата'
        years_sheet['D1'] = f'Год'
        years_sheet['E1'] = 'Количество вакансий'
        column_cells = [years_sheet["A1"], years_sheet['B1'], years_sheet['D1'], years_sheet['E1']]
        for cell in column_cells:
            cell.border = self.border
            cell.font = self.font
        for x in range(len(salaries)):
            years_sheet[f'A{x + 2}'] = list(salaries.keys())[x]
            years_sheet[f'B{x + 2}'] = list(salaries.values())[x]
            years_sheet[f'D{x + 2}'] = list(salaries.keys())[x]
            years_sheet[f'E{x + 2}'] = list(counts.values())[x]
            years_sheet[f'A{x + 2}'].border = self.border
            years_sheet[f'B{x + 2}'].border = self.border
            years_sheet[f'D{x + 2}'].border = self.border
            years_sheet[f'E{x + 2}'].border = self.border
        dim_holder = DimensionHolder(worksheet=years_sheet)

        for col in range(years_sheet.min_column, years_sheet.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(years_sheet, min=col, max=col, width=20)

        years_sheet.column_dimensions = dim_holder

        return years_sheet

    def generate_areas_sheet(self, wb: Workbook, salaries: dict, counts: dict) -> worksheet.Worksheet:
        """
        Генирирует лист в excel-книге со статистикой по городам
        :param wb: excel-книга
        :type wb: Workbook
        :param salaries:словарь средних зарплат по городам
        :type salaries: dict
        :param counts: словарь долей от общего количества вакансий по городам
        :type counts: dict
        :return: excel-лист со статистикой по городам
        """
        areas_sheet = wb.create_sheet("Статистика по городам")
        column_cells = [areas_sheet['A1'], areas_sheet['B1'], areas_sheet['D1'], areas_sheet['E1']]
        for cell in column_cells:
            cell.border = self.border
            cell.font = self.font
        areas_sheet['A1'] = "Город"
        areas_sheet['B1'] = "Уровень зарплат"
        areas_sheet['D1'] = "Город"
        areas_sheet['E1'] = "Доля вакансий"
        for x in range(len(salaries)):
            areas_sheet[f'A{x + 2}'] = list(salaries.keys())[x]
            areas_sheet[f'B{x + 2}'] = list(salaries.values())[x]
            areas_sheet[f'D{x + 2}'] = list(counts.keys())[x]
            areas_sheet[f'E{x + 2}'] = list(counts.values())[x]
            areas_sheet[f'A{x + 2}'].border = self.border
            areas_sheet[f'B{x + 2}'].border = self.border
            areas_sheet[f'D{x + 2}'].border = self.border
            areas_sheet[f'E{x + 2}'].border = self.border
        dim_holder = DimensionHolder(worksheet=areas_sheet)

        for col in range(areas_sheet.min_column, areas_sheet.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(areas_sheet, min=col, max=col, width=20)

        areas_sheet.column_dimensions = dim_holder
        return areas_sheet


data = MultiprocessorHandler().result_data
report = ExcelReport(Side(style="thin", color="000000"), Font(bold=True))
report.generate_excel(data)

